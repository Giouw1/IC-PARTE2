import pandas as pd
import subprocess
import cplex
import networkx as nx
import time
import os
import readershappyinstances
from readershappyinstances import GraphInstance
from docplex.cp.model import CpoModel
from docplex.mp.model import Model
from docplex.mp.relax_linear import LinearRelaxer
import math
import logging
import openpyxl 
import sys
import platform
from contextlib import contextmanager
#Linhas para alteração: 54 : Path do cpoptimizer
#Best bound ta sendo registrado
#306: Não estava conseguindo chamar o código main de c++ pelo terminal:
#A lógica é que eu chamo pelo terminal, e ele gera um arquivo de texto com os resultados, que depois eu leio aqui.
#Folder path mudar caso esteja dando algum problema.
#Construo várias vezes arquivos ou pastas no diretório da função: caso dê problema:
#coloca o caminho completo
# Função para configurar o logger
def setup_logger(name, log_file, level=logging.INFO):
    """Configura o logger para gravar mensagens em um arquivo específico."""
    logger = logging.getLogger(name)
    handler = logging.FileHandler(log_file)
    handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger

# Context Manager para redirecionar stdout e stderr para um arquivo de log
@contextmanager
def redirect_output(log_file):
    """Context manager para redirecionar stdout e stderr para um arquivo de log."""
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    with open(log_file, 'w') as log:
        sys.stdout = log
        sys.stderr = log
        try:
            yield
        finally:
            sys.stdout = original_stdout
            sys.stderr = original_stderr

def maxhs_cp(graph, k, log_file):#comparar
    cpomdl = CpoModel()
    n = len(graph)
    x = cpomdl.binary_var_list(n, name="x")
    objective = cpomdl.sum(
        cpomdl.min([x[j] for j in (graph[i] + [i])])  # mínimo entre os vizinhos de i + i
        for i in range(n)
    )
    cpomdl.maximize(objective)
    cpomdl.add_constraint(cpomdl.sum(x) == k)

    # cpoptimizer_path = '\\opt\\ibm\\ILOG\\CPLEX_Studio2211\\cpoptimizer\\bin\\x86-64_linux\\cpoptimizer.exe'

    start_time = time.time()  

    # Criando o logger para capturar as saídas no log
    logger = setup_logger('CPOptimizer', log_file)

    # Abrindo o arquivo de log corretamente
    with open(log_file, 'w') as log_output:
        logger.info("Iniciando a otimização com CP Optimizer.")

        # Passando o objeto de arquivo log_output para o parâmetro log_output
        solution = cpomdl.solve(
            TimeLimit=600, 
            Presolve='Off', 
            Workers=1, 
           # execfile=cpoptimizer_path, 
            log_output=log_output,
            OptimalityTolerance=0.01 
        )
    
    exec_time = time.time() - start_time  # Tempo de execução

    if solution:
        selected_vertices = [i for i in range(n) if solution[x[i]] == 1]
        happy_vertices = []
        for i in range(n):
            closed_neighbors = graph[i] + [i]  # Vizinhos de i, incluindo o próprio vértice
            if all(solution[x[j]] == 1 for j in closed_neighbors):
                happy_vertices.append(i)
        objective_value = solution.get_objective_value()  # Valor da função objetivo
        gap_cp = solution.get_objective_gap()
        processed_nodes_cp = solution.get_info("NumberOfChoicePoints") #usar em todos
        return selected_vertices, happy_vertices, exec_time, objective_value, gap_cp,processed_nodes_cp
    else:
        logger.error("Solução não encontrada!")
        return None, None, exec_time, None

# def solve_original_model(lista_viz, k, log_file):#comparar
#     mdl = Model(log_output=True)
#     n = len(lista_viz)
#     y = mdl.binary_var_list(n, name="y")
#     h = mdl.continuous_var_list(n, name="h")

#     mdl.maximize(mdl.sum(h[i] for i in range(n)))
#     mdl.add_constraint(mdl.sum(y) == k)
#     for i in range(n):
#         mdl.add_constraint(h[i] <= y[i])
#     for i in range(n):
#         for j in lista_viz[i]:  # j ∈ Γ(i)
#             mdl.add_constraint(h[i] <= y[j])
#     for i in range(n):
#         mdl.add_constraint(h[i] >= mdl.sum(y[j] for j in lista_viz[i]) - len(lista_viz[i]) + y[i])

#     mdl.parameters.mip.strategy.heuristicfreq.set(-1)
#     mdl.parameters.mip.cuts.mircut.set(-1)
#     mdl.parameters.mip.cuts.implied.set(-1)
#     mdl.parameters.mip.cuts.gomory.set(-1)
#     mdl.parameters.mip.cuts.flowcovers.set(-1)
#     mdl.parameters.mip.cuts.pathcut.set(-1)
#     mdl.parameters.mip.cuts.liftproj.set(-1)
#     mdl.parameters.mip.cuts.zerohalfcut.set(-1)
#     mdl.parameters.mip.cuts.cliques.set(-1)
#     mdl.parameters.mip.cuts.covers.set(-1)
#     mdl.parameters.threads.set(1)
#     mdl.parameters.clocktype.set(1)
#     mdl.parameters.timelimit.set(600)
#     mdl.parameters.preprocessing.presolve.set(0)
#     mdl.parameters.preprocessing.boundstrength.set(0)
#     mdl.parameters.preprocessing.coeffreduce.set(0)
#     mdl.parameters.preprocessing.relax.set(0)
#     mdl.parameters.preprocessing.aggregator.set(0)
#     mdl.parameters.preprocessing.reduce.set(0)
#     mdl.parameters.preprocessing.reformulations.set(0)

#     # Criando o logger para capturar as saídas no log
#     logger = setup_logger('OriginalModel', log_file)
#     with redirect_output(log_file):
#         logger.info("Iniciando a otimização com o modelo original.")

#         start_time = time.time()  # Tempo de início da execução
#         solution = mdl.solve()
#         exec_time = time.time() - start_time  # Tempo de execução

#     if solution:
#         selected_vertices = [i for i in range(n) if solution.get_value(y[i]) == 1]
#         happy_vertices = [i for i in range(n) if solution.get_value(h[i]) == 1]
#         objective_value = solution.objective_value  # Valor da função objetivo
#         gap = mdl.solve_details.mip_relative_gap
#         processed_nodes = mdl.solve_details.nb_nodes_processed #usar em todos
#         return selected_vertices, happy_vertices, exec_time, objective_value,gap,processed_nodes
#     else:
#         logger.error("Solução não encontrada!")
#         return None, None, exec_time, None


def solve_original_model(lista_viz, k, log_file, log_lp_file):
    mdl = Model(log_output=True)
    n = len(lista_viz)
    y = mdl.binary_var_list(n, name="y")
    h = mdl.continuous_var_list(n, name="h")

    mdl.maximize(mdl.sum(h[i] for i in range(n)))
    mdl.add_constraint(mdl.sum(y) == k)
    for i in range(n):
        mdl.add_constraint(h[i] <= y[i])
    for i in range(n):
        for j in lista_viz[i]:
            mdl.add_constraint(h[i] <= y[j])
    for i in range(n):
        mdl.add_constraint(h[i] >= mdl.sum(y[j] for j in lista_viz[i]) - len(lista_viz[i]) + y[i])

    # Parâmetros do solver
    mdl.parameters.mip.strategy.heuristicfreq.set(-1)
    mdl.parameters.mip.cuts.mircut.set(-1)
    mdl.parameters.mip.cuts.implied.set(-1)
    mdl.parameters.mip.cuts.gomory.set(-1)
    mdl.parameters.mip.cuts.flowcovers.set(-1)
    mdl.parameters.mip.cuts.pathcut.set(-1)
    mdl.parameters.mip.cuts.liftproj.set(-1)
    mdl.parameters.mip.cuts.zerohalfcut.set(-1)
    mdl.parameters.mip.cuts.cliques.set(-1)
    mdl.parameters.mip.cuts.covers.set(-1)
    mdl.parameters.threads.set(1)
    mdl.parameters.clocktype.set(1)
    mdl.parameters.timelimit.set(60)
    mdl.parameters.preprocessing.presolve.set(0)
    mdl.parameters.preprocessing.boundstrength.set(0)
    mdl.parameters.preprocessing.coeffreduce.set(0)
    mdl.parameters.preprocessing.relax.set(0)
    mdl.parameters.preprocessing.aggregator.set(0)
    mdl.parameters.preprocessing.reduce.set(0)
    mdl.parameters.preprocessing.reformulations.set(0)

    logger = setup_logger('OriginalModel', log_file)
    with redirect_output(log_file):
        logger.info("Iniciando a otimização com o modelo original.")
        start_time = time.time()
        solution = mdl.solve()
        exec_time = time.time() - start_time

    if solution:
        selected_vertices = [i for i in range(n) if solution.get_value(y[i]) == 1]
        happy_vertices = [i for i in range(n) if solution.get_value(h[i]) == 1]
        objective_value = solution.objective_value
        gap = mdl.solve_details.mip_relative_gap
        best_bound = solution.solve_details.best_bound
        processed_nodes = mdl.solve_details.nb_nodes_processed #usar em todos
    else:
        logger.error("Solução não encontrada!")
        return None, None, 0, None, None

    # Resolver a LP
    LP = LinearRelaxer.make_relaxed_model(mdl, name="RelaxedModel")
    logger_lp = setup_logger('RelaxedModel', log_lp_file)
    with redirect_output(log_lp_file):
        logger_lp.info("Resolvendo a relaxação linear (LP).")
        start_lp_time = time.time()
        lp_solution = LP.solve()
        lp_exec_time = time.time() - start_lp_time

        if lp_solution:
            lp_objective_value = lp_solution.objective_value
            # Recupera variáveis do modelo LP
            lp_y_vars = [LP.get_var_by_name(y[i].name) for i in range(n)]
            lp_h_vars = [LP.get_var_by_name(h[i].name) for i in range(n)]

            lp_y_values = [lp_solution.get_value(var) for var in lp_y_vars]
            lp_h_values = [lp_solution.get_value(var) for var in lp_h_vars]


            logger_lp.info(f"Tempo de execução da LP: {lp_exec_time:.4f} segundos")
            logger_lp.info(f"Valor da função objetivo da LP: {lp_objective_value:.4f}")
            logger_lp.info("Valores de y (LP): " + str([round(val, 4) for val in lp_y_values]))
            logger_lp.info("Valores de h (LP): " + str([round(val, 4) for val in lp_h_values]))
        else:
            logger_lp.error("Falha ao resolver o modelo relaxado (LP).")

    return selected_vertices, happy_vertices, exec_time, objective_value, gap, processed_nodes, best_bound 


# Função para salvar resultados no Excel na primeira linha disponível
def save_results_to_excel(results, log_folder, filename="teste_rfael_best_bound.xlsx"):
    excel_file_path = os.path.join(log_folder, filename)

    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        row = sheet.max_row + 1
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([  "instance_name", "k", "selected_vertices_cp", "happy_vertices_cp", 
                        "exec_time_cp", "objective_value_cp", "gap_cp","processed_nodes_cp",
                        "selected_vertices_orig", "happy_vertices_orig", 
                        "exec_time_orig", "objective_value_orig","best_bound","gap","processed_nodes","happy_vertices_back",
                        "selected_vertices_back","objective_value_back","processed_nodes_back",
                        "tempo_back","validaback","best_objective_value_heur", "all_objective_value_heur", "total_time_heur",
                        "happy_vertices_heur","colored_vertices_heur","validaheur","gamma"
                     ])
        row = 2

    for result in results:
        # Converte as listas para strings
        selected_vertices_cp_str = result["selected_vertices_cp"]
        happy_vertices_cp_str = result["happy_vertices_cp"]
        selected_vertices_orig_str = result["selected_vertices_orig"]
        happy_vertices_orig_str = result["happy_vertices_orig"]


        # happy_vertices_cp_str = ','.join(map(str, result["happy_vertices_cp"]))
        # selected_vertices_orig_str = ','.join(map(str, result["selected_vertices_orig"]))







        # happy_vertices_orig_str = ','.join(map(str, result["happy_vertices_orig"]))
        # selected_vertices_ben_aut_str= ','.join(map(str, result["selected_vertices_ben_aut"]))
        # happy_vertices_ben_aut_str= ','.join(map(str, result["happy_vertices_ben_aut"]))


        sheet.append([result["instance_name"], result["k"], selected_vertices_cp_str,
                      happy_vertices_cp_str, result["exec_time_cp"], result["objective_value_cp"], result["gap_cp"],result["processed_nodes_cp"],
                      selected_vertices_orig_str, happy_vertices_orig_str,
                      result["exec_time_orig"], result["objective_value_orig"], result["best_bound"], result["gap"], result["processed_nodes"],
                      result["happy_vertices_back"], result["selected_vertices_back"], result["objective_value_back"], 
                      result["processed_nodes_back"], result["tempo_back"], result["validaback"],result["best_objective_value_heur"],
                      result["all_objective_value_heur"], result["total_time_heur"], result["happy_heur"], result["colored_heur"],
                      result["validaheur"], result["gamma"]
                    ])
        
        

    workbook.save(excel_file_path)
    print(f"Resultados salvos em {excel_file_path}")


# Função para criar a pasta de logs
def create_log_folder():
    log_folder = 'pasta/logs'
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)  # Cria a pasta se não existir
    return log_folder

# Função para calcular os valores de k a serem testados
def get_k_values(num_vertices):
    k_values = [
        math.ceil(0.40 * num_vertices), 
        math.ceil(0.50 * num_vertices),
        math.ceil(0.60 * num_vertices)
    ]
    if (num_vertices<=10):
        k_values = [
            math.ceil(0.80 * num_vertices),
            math.ceil(0.60 * num_vertices)
        ]
    # Filtrar os valores para manter apenas os maiores que o valor de k
    #k_values = [value for value in k_values if value > k]

    # Adiciona k na lista, caso seja maior que 0 e não tenha sido descartado
   # if k > 0 and k not in k_values:
   #     k_values.insert(0, k)  

    return sorted(k_values)
# Função para obter arquivos em uma pasta com uma extensão específica
def get_files_in_folder(folder_path, extension=".txt"):
    files = [f for f in os.listdir(folder_path) if f.endswith(extension)]
    full_paths = [os.path.join(folder_path, f) for f in files]
    return full_paths

# Função principal para processar os arquivos
def process_files_in_folder(folder_path):
    files = get_files_in_folder(folder_path)
    results = []
    # Criar a pasta de logs, caso não exista
    log_folder = create_log_folder() 
    for file in files:
        print(f"\nProcessando arquivo: {file}")
        instancia = GraphInstance(file)
        adj_list, precolored, vertices, arestas,kk = instancia.get_data()
        kvalues = get_k_values(len(adj_list))
        limit = 600#passar isso como parametro da file?
        seed = 10#passar isso como parametro da file?
        n = 50
        gamma = 1.2
        for k in kvalues:
            subprocess.run(['codigoscpp/scripttodo.exe', file, str(k), str(limit), str(seed),str(n), str(gamma)])
            # Nome da instância (arquivo sem extensão)
            instance_name = os.path.basename(file)
            nodes_processed,time_spent,answer_back, happy_back,colored_back,answer_heur, answers_heur, time_heur, happy_heur, colored_heur,validaheur,validaback = read_data_giovanni("resultcpp.txt")
            
            # nodes_processed = 0
            # time_spent = 0
            # answer_back = 0
            # colored_back = []
            # happy_back = []
            # answer_heur = 0
            # answers_heur = []
            # time_heur = 0
            # happy_heur = []
            # colored_heur = []

            # validaheur = 0
            
            # Calcular os valores de k a serem testados

            # Resolver o problema para os valores de k calculados
            print(f"\nValor de k: {k}")
            
            # Definir o nome do arquivo de log
            log_file_path = os.path.join(log_folder, f'log_{instance_name}_k{k}.txt')
            log_lp_file_path = os.path.join(log_folder, f'log_{instance_name}_k{k}.txt')
            
            #Algoritmo CP Optimizer
            selected_vertices_cp, happy_vertices_cp, exec_time_cp, objective_value_cp, gap_cp,processed_nodes_cp = maxhs_cp(adj_list, k ,log_file_path)
            print("Algoritmo CP Optimizer:")
            print("Vértices selecionados:", selected_vertices_cp)
            print("Vértices felizes:", happy_vertices_cp)
            print(f"Tempo de execução: {exec_time_cp:.4f} segundos")
            print(f"Valor da função objetivo: {objective_value_cp}")
            print(f"Número de nós explorados: {processed_nodes_cp}")
            print(f"Valor do GAP do CP: {gap_cp}")


            # selected_vertices_cp = []
            # happy_vertices_cp = []
            # exec_time_cp = 0
            # objective_value_cp = 0
            # gap_cp = 0
            # processed_nodes_cp = 0
            # Algoritmo Original
            selected_vertices_orig, happy_vertices_orig, exec_time_orig, objective_value_orig,gap,processed_nodes,best_bound = solve_original_model(adj_list, k ,log_file_path,log_lp_file_path)
            # selected_vertices_orig = []
            # happy_vertices_orig = []
            # exec_time_orig = 0
            # objective_value_orig = 0
            # gap = 0
            # processed_nodes = 0
            # print("Algoritmo Original:")
            # print("Vértices selecionados:", selected_vertices_orig)
            # print("Vértices felizes:", happy_vertices_orig)
            # print(f"Tempo de execução: {exec_time_orig:.4f} segundos")
            # print(f"Valor da função objetivo: {objective_value_orig}")
            # print(f"Número de nós explorados: {processed_nodes}")
            # print(f"Valor do GAP: {gap}")
            #aqui entra os códigos

                # Benders
            # selected_vertices_ben_aut, happy_vertices_ben_aut, exec_time_ben_aut, objective_value_ben_aut = model_benders(adj_list, k_val)
            # print("Algoritmo Automatico de Benders:")
            # print("Vértices selecionados:", selected_vertices_ben_aut)
            # print("Vértices felizes:", happy_vertices_ben_aut)
            # print(f"Tempo de execução: {exec_time_ben_aut:.4f} segundos")
    #            print(f"Valor da função objetivo: {objective_value_ben_aut}")
        

            #Armazenando os resultados como listas (strings)
            #nesse results, introduzir os valores da heurística e do código backtracking
            results.append({
                "instance_name": instance_name,
                "k": k,
                "selected_vertices_cp": str(selected_vertices_cp),
                "happy_vertices_cp": str(happy_vertices_cp),
                "exec_time_cp": exec_time_cp,
                "objective_value_cp": objective_value_cp,
                "gap_cp":gap_cp,
                "processed_nodes_cp":processed_nodes_cp,
                "selected_vertices_orig": str(selected_vertices_orig),
                "happy_vertices_orig": str(happy_vertices_orig),
                "exec_time_orig": exec_time_orig,
                "objective_value_orig": objective_value_orig,
                "best_bound":best_bound,
                "gap": gap,
                "processed_nodes":processed_nodes,
                "happy_vertices_back": str(happy_back),
                "selected_vertices_back": str(colored_back),
                "objective_value_back" : answer_back,
                "processed_nodes_back": nodes_processed,
                "tempo_back": time_spent,
                "validaback": validaback,
                "best_objective_value_heur": answer_heur,
                "all_objective_value_heur": str(answers_heur),
                "total_time_heur": time_heur,
                "happy_heur": str(happy_heur),
                "colored_heur": str(colored_heur),
                "validaheur": validaheur,
                "gamma": gamma


                # "selected_vertices_ben_aut": str(selected_vertices_ben_aut),
                # "happy_vertices_ben_aut": str(happy_vertices_ben_aut),
                # "exec_time_ben_aut": exec_time_ben_aut,
                # "objective_value_ben_aut": objective_value_ben_aut

            })
                    #coloca em ja usados, o arquivo
            # Salvar os resultados em um arquivo Excel dentro da pasta CP
            save_results_to_excel(results, log_folder)
            results = []
#       os.rename(file,f"pasta\\ja_usadas\\{file[6:]}")
        destino = os.path.join("pasta/ja_usadas", os.path.basename(file))
        try:
            os.rename(file, destino)
        except FileNotFoundError:
            print(f"Erro ao mover o arquivo: {file} não encontrado.")


def read_data_giovanni(path):
    nodes_processed = 0
    time_spent = 0
    answer_back = 0
    colored_back = []
    happy_back = []
    answer_heur = 0
    answers_heur = []
    marker = 0
    tempo_heur = 0
    happy_heur = []
    colored_heur = []
    validaheur = 1
    validaback = 1
    with open(path, 'r') as file:
        first_line = file.readline()
        for i in range(len(first_line)):
            if (first_line[i] == " "):
                for j in range(i+1,len(first_line)):
                    if (first_line[j] == " "):
                        nodes_processed = int(first_line[0:i])
                        time_spent = float(first_line[i+1:j])
                        answer_back = int(first_line[j+1:])
        for line in file:
            if line.strip() == "ERROH":
                validaheur = 0
                continue
            if line.strip() == "OKH":
                validaheur = 1
                continue
            if line.strip() == "ERROB":
                validaback= 0
                continue
            if line.strip() == "OKB":
                validaback = 1
                continue
            if (line[0] == "h" and marker == 0):
                happy_back+=[int(line[2:].strip())]
            if (line[0] == "c" and marker == 0):
                colored_back+=[int(line[2:].strip())]
            if (line[0] == "h" and marker == 1):
                happy_heur+=[int(line[2:].strip())]
            if (line[0] == "c" and marker == 1):
                colored_heur+=[int(line[2:].strip())]
            if (line[0] == "r"):
                answers_heur+= [int(line[2:].strip())]
            if (line[0] != "h" and line[0] != "c" and line[0]!= "r"  and marker == 1):
                tempo_heur= float(line.strip())
            if (line[0] != "h" and line[0] != "c" and line[0]!= "r" and (marker == 0)):
                marker = 1
                answer_heur = int(line.strip())
        
    file.close()
    return (nodes_processed,time_spent,answer_back, happy_back,colored_back,answer_heur,answers_heur,tempo_heur, happy_heur,colored_heur,validaheur,validaback)
# Caminho da pasta com os arquivos
#folder_path = '/home/rafael/Documents/HappySet/MIHS/inputs/happygen/output/testes/testes_cp_benders'
# folder_path ='/home/rafael/Documents/HappySet/MIHS/inputs/happygen/output/testes/7-10'
# Chama a função para processar os arquivos da pasta
#process_files_in_folder(folder_path)
# Caminho da pasta com os arquivos
folder_path = "7-10"
# folder_path ='/home/rafael/Documents/HappySet/MIHS/inputs/happygen/output/testes/7-10'

# Chama a função para processar os arquivos da pasta
process_files_in_folder(folder_path)    
#processed nodes cp
#heuristica ver tudo
